#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
import xml.etree.ElementTree as ET
from collections import Counter
from dataclasses import asdict, dataclass, fields
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import fitz
from openpyxl import Workbook

from batch_paper_fetch import discover_pdf_url, get_json, slugify, try_download


class ChineseArgumentParser(argparse.ArgumentParser):
    def format_help(self) -> str:
        text = super().format_help()
        replacements = {
            "usage: ": "用法：",
            "options:\n": "可选参数：\n",
            "positional arguments:\n": "位置参数：\n",
            "show this help message and exit": "显示这条帮助信息并退出",
        }
        for source, target in replacements.items():
            text = text.replace(source, target)
        return text


DEFAULT_CONFIG: dict[str, Any] = {
    "project_name": "扎根文献每日监测",
    "queries": ["创业 即兴行为 扎根理论", "创业 资源视角 扎根理论"],
    "sources": ["local", "openalex", "arxiv", "semantic_scholar"],
    "days_back": 30,
    "max_results_per_query": 20,
    "download_pdfs": True,
    "max_pdf_pages_for_coding": 30,
    "sleep_seconds": 0.4,
    "outdir": "output/grounded_monitor",
    "baseline_paths": [],
    "covered_topics": [],
    "local_library_paths": [],
    "max_local_papers": 20,
    "translation": {
        "enabled": False,
        "provider": "openai_compatible",
        "model": "",
        "api_url": "",
        "api_key_env": "GROUNDED_TRANSLATE_API_KEY",
    },
    "skill_dirs": [],
    "agent": {
        "enabled": True,
        "max_turns": 12,
        "trace_candidates": False,
    },
    "assistant": {
        "enabled": False,
        "provider": "openai_compatible",
        "model": "",
        "api_url": "",
        "api_key_env": "GROUNDED_AGENT_API_KEY",
        "temperature": 0.2,
        "max_context_rows": 10,
        "max_context_chars": 24000,
        "answer_dirname": "qa_answers",
        "report_dirname": "industry_reports",
    },
    "context_compression": {
        "enabled": True,
        "char_threshold": 30000,
        "max_recent_rows": 12,
    },
}


BUILTIN_SKILLS = [
    {
        "name": "multi-source-literature-scout",
        "goal": "按优先级从本地论文库、OpenAlex、arXiv、Semantic Scholar 等多源增量检索文献，并统一去重。",
    },
    {
        "name": "grounded-open-coder",
        "goal": "依据全文或摘要中的证据句，对文献进行开放编码，优先保留前因、结果、机制、边界条件和未来研究线索。",
    },
    {
        "name": "axial-theme-cluster",
        "goal": "把开放编码聚合为前因、结果、机制、边界条件等聚焦编码，并形成一句话的主轴概括。",
    },
    {
        "name": "gap-alert-auditor",
        "goal": "将新文献主题与用户既有论文或已覆盖主题进行对比，标记新增主题并给出补写方向提醒。",
    },
    {
        "name": "state-memory-keeper",
        "goal": "维护运行状态、搜索历史、下载历史和主题记忆，使长周期任务可断点续跑并保留研究脉络。",
    },
    {
        "name": "skill-discovery-router",
        "goal": "自动发现外部 SKILL.md 目录，基于查询语义挑选与当前研究任务最相关的 skills 作为辅助能力。",
    },
    {
        "name": "agent-runtime-loop",
        "goal": "将检索、编码、对比、记忆压缩拆成显式的执行步骤，并把每一步写入执行轨迹，形成可恢复的轻量 agent loop。",
    },
    {
        "name": "context-compressor",
        "goal": "当历史文献和运行上下文变长时，自动生成紧凑摘要，保留目标、发现、已完成事项和下一步建议。",
    },
    {
        "name": "evidence-rag-answerer",
        "goal": "围绕用户问题从文献表、摘要、证据句和本地 PDF 中拼装证据上下文，并生成带来源说明的回答。",
    },
    {
        "name": "industry-report-writer",
        "goal": "围绕某个行业主题自动汇总关键文献证据，生成执行摘要、驱动因素、竞争格局、风险与建议等报告章节。",
    },
]


PROMPT_TEMPLATES = {
    "agent_planner": (
        "先发现相关 skills，再基于查询与数据源生成明确执行计划：检索、去重、编码、缺口提醒、记忆压缩。"
    ),
    "source_router": (
        "先查本地论文库和历史记录，再查外部来源。多源结果统一去重后再进入编码阶段。"
    ),
    "open_coding": (
        "你是一名使用扎根理论做文献编码的研究助手。请仅依据输入文本，优先抽取与研究主题直接相关的原始概念。"
        "保留作者原文术语，避免过早抽象。重点识别前因、结果、机制、边界条件、研究不足和未来研究。"
    ),
    "axial_coding": (
        "请将开放编码归类为聚焦编码，优先按前因、结果、机制、边界条件、研究对象、方法与理论基础组织。"
        "输出时明确类属之间的可能关系。"
    ),
    "selective_summary": (
        "请用一句话概括该文主线：在什么情境下，什么前因通过何种机制影响什么结果，并受到哪些边界条件约束。"
    ),
    "gap_alert": (
        "请将新文献的主题编码与用户已覆盖主题做对比，若存在新增主题、变量关系或未来研究方向，请单独提醒。"
        "提醒内容需可直接转化为论文补写方向。"
    ),
    "qa_answer": (
        "你是一名研究型行业分析助手。请优先依据给定证据回答问题，先给直接答案，再列关键证据、局限和下一步建议。"
        "若证据不足要明确说明，不要编造。"
    ),
    "industry_report": (
        "你是一名行业研究员。请结合给定文献与证据，生成结构化行业报告，至少覆盖执行摘要、行业现状、关键驱动因素、"
        "竞争格局、风险与不确定性、机会与建议、后续跟踪指标。"
    ),
}


CODEBOOK = {
    "antecedents": {
        "环境不确定性": ["不确定", "动态性", "动荡", "vuca", "uncertainty", "turbulence", "hostile", "dynamism"],
        "资源约束与资源拼凑": ["资源", "资源拼凑", "resource", "bricolage", "scarcity", "constraint", "resource orchestration"],
        "创业者认知与经验": ["认知", "经验", "创业者", "高管", "manager", "cognition", "experience", "upper echelons"],
        "网络与关系嵌入": ["网络", "关系", "信任", "社会资本", "network", "tie", "trust", "embeddedness"],
        "组织能力与战略导向": ["能力", "导向", "动态能力", "吸收能力", "创业导向", "capability", "orientation", "absorptive"],
        "制度与技术环境": ["制度", "政策", "监管", "技术环境", "toe", "institutional", "policy", "regulation"],
    },
    "outcomes": {
        "企业绩效与成长": ["绩效", "成长", "盈利", "performance", "growth", "survival"],
        "创新与机会开发": ["创新", "机会", "新产品", "innovation", "opportunity", "new product"],
        "竞争优势与合法性": ["竞争优势", "合法性", "声誉", "竞争力", "advantage", "legitimacy", "reputation"],
        "韧性与适应": ["韧性", "适应", "恢复", "resilience", "adaptation", "recovery"],
        "实施与采纳成效": ["采纳成效", "实施成效", "adoption outcome", "implementation outcome", "使用效果"],
    },
    "mechanisms": {
        "资源整合与编排": ["整合", "编排", "重构", "resource integration", "resource orchestration", "reconfigure"],
        "知识学习与吸收": ["学习", "知识", "吸收", "knowledge", "learning", "absorptive"],
        "认知加工与决策逻辑": ["决策", "逻辑", "注意力", "信息处理", "decision", "attention", "sensemaking"],
        "协同与关系转化": ["协同", "合作", "转化", "联结", "collaboration", "conversion", "coordination"],
    },
    "boundaries": {
        "企业阶段与生命周期": ["阶段", "生命周期", "成长阶段", "stage", "life cycle", "lifecycle"],
        "团队与个体异质性": ["异质性", "团队", "个体差异", "heterogeneity", "team", "individual"],
        "环境动态性与行业情境": ["情境", "行业", "动态性", "环境", "context", "industry", "environmental dynamism"],
        "组织规模与区域条件": ["规模", "区域", "所有制", "institution type", "size", "region", "ownership"],
    },
    "future_directions": {
        "未来研究与理论延展": ["未来研究", "研究展望", "不足", "future research", "limitation", "extension"],
        "方法与数据改进": ["纵向", "实验", "多案例", "panel", "longitudinal", "experiment", "multi-case"],
        "新情境拓展": ["新情境", "跨层次", "数字化", "平台", "genai", "aigc", "cross-level"],
    },
    "research_objects": {
        "个体与创业者层": ["创业者", "管理者", "创始人", "individual", "entrepreneur", "founder", "manager"],
        "团队与组织层": ["团队", "组织", "企业", "team", "organization", "firm", "venture"],
        "生态与平台层": ["生态", "平台", "网络", "ecosystem", "platform", "network"],
        "行业与制度层": ["行业", "产业", "制度", "区域", "industry", "sector", "institution", "region"],
    },
    "methods": {
        "案例与扎根方法": ["案例", "扎根", "案例研究", "case study", "grounded theory", "multiple case"],
        "实证与量化方法": ["回归", "问卷", "结构方程", "面板", "regression", "survey", "sem", "panel"],
        "文本与二手数据": ["文本", "档案", "二手数据", "文本分析", "text analysis", "archival", "secondary data"],
        "混合方法": ["混合方法", "mixed method", "triangulation", "多方法"],
    },
    "theory_basis": {
        "资源基础与能力理论": ["资源基础", "动态能力", "resource-based", "rbv", "dynamic capability"],
        "制度与情境理论": ["制度理论", "制度逻辑", "institutional", "contingency", "contextual"],
        "认知与行为理论": ["认知", "sensemaking", "行为理论", "attention", "behavioral"],
        "网络与社会资本理论": ["社会资本", "嵌入", "network theory", "social capital", "embeddedness"],
    },
}


CODEBOOK_LOOKUP = {
    label: keywords
    for labels in CODEBOOK.values()
    for label, keywords in labels.items()
}


SOURCE_COLUMNS = [
    "run_date",
    "query",
    "source_name",
    "source_type",
    "project_name",
    "external_id",
    "doi",
    "title",
    "title_zh",
    "authors",
    "journal",
    "year",
    "publication_date",
    "language",
    "cited_by_count",
    "peer_reviewed",
    "primary_topic",
    "topics",
    "abstract",
    "abstract_zh",
    "source_url",
    "pdf_url",
    "pdf_downloaded",
    "local_pdf",
    "open_codes",
    "axial_summary",
    "selective_summary",
    "antecedents",
    "outcomes",
    "mechanisms",
    "boundaries",
    "future_directions",
    "research_objects",
    "methods",
    "theory_basis",
    "novel_themes",
    "novel_relations",
    "gap_focus",
    "alert_level",
    "recommendation",
    "hypotheses_propositions",
    "independent_vars",
    "mediator_moderator_vars",
    "dependent_vars",
    "control_vars",
    "future_research_directions",
    "future_direction_codes",
    "open_code_details",
    "axial_relations",
    "selective_proposition",
    "coding_confidence",
    "evidence_sentences",
]


@dataclass
class SearchCandidate:
    query: str
    source_name: str
    source_type: str
    external_id: str
    doi: str
    title: str
    authors: str
    journal: str
    year: str
    publication_date: str
    language: str
    cited_by_count: int
    peer_reviewed: str
    primary_topic: str
    topics: str
    abstract: str
    source_url: str
    pdf_url: str
    local_pdf: str
    preview_text: str


@dataclass
class MonitorRow:
    run_date: str = ""
    query: str = ""
    source_name: str = ""
    source_type: str = ""
    project_name: str = ""
    external_id: str = ""
    doi: str = ""
    title: str = ""
    title_zh: str = ""
    authors: str = ""
    journal: str = ""
    year: str = ""
    publication_date: str = ""
    language: str = ""
    cited_by_count: int = 0
    peer_reviewed: str = ""
    primary_topic: str = ""
    topics: str = ""
    abstract: str = ""
    abstract_zh: str = ""
    source_url: str = ""
    pdf_url: str = ""
    pdf_downloaded: str = "否"
    local_pdf: str = ""
    open_codes: str = ""
    axial_summary: str = ""
    selective_summary: str = ""
    antecedents: str = ""
    outcomes: str = ""
    mechanisms: str = ""
    boundaries: str = ""
    future_directions: str = ""
    research_objects: str = ""
    methods: str = ""
    theory_basis: str = ""
    novel_themes: str = ""
    novel_relations: str = ""
    gap_focus: str = ""
    alert_level: str = ""
    recommendation: str = ""
    hypotheses_propositions: str = ""
    independent_vars: str = ""
    mediator_moderator_vars: str = ""
    dependent_vars: str = ""
    control_vars: str = ""
    future_research_directions: str = ""
    future_direction_codes: str = ""
    open_code_details: str = ""
    axial_relations: str = ""
    selective_proposition: str = ""
    coding_confidence: str = ""
    evidence_sentences: str = ""


def normalize_text(text: str) -> str:
    text = text.replace("\x00", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip().lower()


def query_tokens(text: str) -> list[str]:
    tokens = re.findall(r"[\u4e00-\u9fffA-Za-z0-9]+", text.lower())
    return [token for token in tokens if len(token) >= 2]


def safe_join(items: list[str]) -> str:
    return "；".join(items)


def candidate_key(title: str, doi: str = "", external_id: str = "") -> str:
    if doi:
        return f"doi:{doi.lower()}"
    if external_id:
        return f"id:{external_id.lower()}"
    normalized = re.sub(r"[^\w\u4e00-\u9fff]+", "", title.casefold())
    return f"title:{normalized}"


def merge_config(user_config: dict[str, Any]) -> dict[str, Any]:
    merged = json.loads(json.dumps(DEFAULT_CONFIG, ensure_ascii=False))
    for key, value in user_config.items():
        if key == "translation" and isinstance(value, dict):
            merged["translation"].update(value)
        elif key in {"agent", "assistant", "context_compression"} and isinstance(value, dict):
            merged[key].update(value)
        else:
            merged[key] = value
    merged["queries"] = [q.strip() for q in merged.get("queries", []) if q.strip()]
    merged["sources"] = [q.strip() for q in merged.get("sources", []) if q.strip()]
    merged["covered_topics"] = [q.strip() for q in merged.get("covered_topics", []) if q.strip()]
    merged["baseline_paths"] = [str(Path(p).expanduser()) for p in merged.get("baseline_paths", []) if str(p).strip()]
    merged["local_library_paths"] = [str(Path(p).expanduser()) for p in merged.get("local_library_paths", []) if str(p).strip()]
    merged["skill_dirs"] = [str(Path(p).expanduser()) for p in merged.get("skill_dirs", []) if str(p).strip()]
    return merged


def invert_abstract(index: dict[str, list[int]] | None) -> str:
    if not index:
        return ""
    size = max((max(pos) for pos in index.values() if pos), default=-1) + 1
    words = [""] * size
    for token, positions in index.items():
        for pos in positions:
            if 0 <= pos < size:
                words[pos] = token
    text = " ".join(word for word in words if word)
    return re.sub(r"\s+([,.;:!?])", r"\1", text).strip()


def read_text_file(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in {".txt", ".md"}:
        return path.read_text(encoding="utf-8", errors="ignore")
    if suffix == ".pdf":
        doc = fitz.open(path)
        return "\n".join(page.get_text() for page in doc)
    if suffix == ".docx":
        try:
            from docx import Document  # type: ignore
        except Exception:
            return ""
        doc = Document(path)
        return "\n".join(p.text for p in doc.paragraphs)
    return ""


def extract_pdf_text(pdf_path: Path, page_limit: int | None = None) -> str:
    doc = fitz.open(pdf_path)
    chunks = []
    total = doc.page_count if page_limit is None else min(page_limit, doc.page_count)
    for i in range(total):
        chunks.append(doc.load_page(i).get_text())
    return "\n".join(chunks)


def extract_sentences(text: str, limit: int = 160) -> list[str]:
    parts = re.split(r"(?<=[。！？.!?;；])\s+|(?<=[。！？.!?;；])", text.replace("\n", " "))
    out: list[str] = []
    seen: set[str] = set()
    for part in parts:
        sentence = re.sub(r"\s+", " ", part).strip(" -:：;；,.，")
        if len(sentence) < 18:
            continue
        if sentence in seen:
            continue
        seen.add(sentence)
        out.append(sentence)
        if len(out) >= limit:
            break
    return out


def match_labels(text: str) -> dict[str, list[str]]:
    normalized = normalize_text(text)
    found: dict[str, list[str]] = {}
    for category, labels in CODEBOOK.items():
        hits: list[str] = []
        for label, keywords in labels.items():
            if any(keyword.lower() in normalized for keyword in keywords):
                hits.append(label)
        found[category] = hits
    return found


def build_open_code_records(text: str, limit: int = 18) -> list[dict[str, str]]:
    sentences = extract_sentences(text, limit=260)
    records: list[dict[str, str]] = []
    seen: set[tuple[str, str, str]] = set()
    for sentence in sentences:
        lowered = sentence.lower()
        for category, labels in CODEBOOK.items():
            for label, keywords in labels.items():
                matched_keyword = ""
                for keyword in keywords:
                    if keyword.lower() in lowered:
                        matched_keyword = keyword
                        break
                if not matched_keyword:
                    continue
                key = (category, label, sentence)
                if key in seen:
                    continue
                seen.add(key)
                records.append(
                    {
                        "category": category,
                        "label": label,
                        "keyword": matched_keyword,
                        "evidence": sentence,
                    }
                )
                if len(records) >= limit:
                    return records
    return records


def aggregate_records(records: list[dict[str, str]]) -> dict[str, list[str]]:
    grouped: dict[str, list[str]] = {category: [] for category in CODEBOOK}
    for record in records:
        category = record["category"]
        label = record["label"]
        if label not in grouped[category]:
            grouped[category].append(label)
    return grouped


def summarize_open_codes(matched: dict[str, list[str]]) -> list[str]:
    codes: list[str] = []
    for category in (
        "antecedents",
        "outcomes",
        "mechanisms",
        "boundaries",
        "future_directions",
        "research_objects",
        "methods",
        "theory_basis",
    ):
        codes.extend(matched.get(category, []))
    return codes


def flatten_labels(matched: dict[str, list[str]]) -> set[str]:
    out: set[str] = set()
    for values in matched.values():
        out.update(values)
    return out


def best_evidence(text: str, matched: dict[str, list[str]], limit: int = 6) -> list[str]:
    sentences = extract_sentences(text, limit=260)
    picked: list[str] = []
    seen: set[str] = set()
    for labels in matched.values():
        for label in labels:
            for keyword in CODEBOOK_LOOKUP[label]:
                for sentence in sentences:
                    if keyword.lower() in sentence.lower() and sentence not in seen:
                        seen.add(sentence)
                        picked.append(sentence)
                        break
                if len(picked) >= limit:
                    return picked
            if len(picked) >= limit:
                return picked
    return picked[:limit]


def build_open_code_details(records: list[dict[str, str]], limit: int = 12) -> str:
    lines: list[str] = []
    for record in records[:limit]:
        category_label = {
            "antecedents": "前因",
            "outcomes": "结果",
            "mechanisms": "机制",
            "boundaries": "边界条件",
            "future_directions": "未来研究",
            "research_objects": "研究对象",
            "methods": "研究方法",
            "theory_basis": "理论基础",
        }.get(record["category"], record["category"])
        lines.append(f"{category_label}|{record['label']}|{record['evidence'][:180]}")
    return " || ".join(lines)


def build_relation_signatures(matched: dict[str, list[str]]) -> list[str]:
    antecedents = matched.get("antecedents", [])[:2]
    mechanisms = matched.get("mechanisms", [])[:2]
    outcomes = matched.get("outcomes", [])[:2]
    boundaries = matched.get("boundaries", [])[:2]
    signatures: list[str] = []
    if antecedents and mechanisms and outcomes:
        for antecedent in antecedents:
            for mechanism in mechanisms:
                for outcome in outcomes:
                    signatures.append(f"{antecedent}->{mechanism}->{outcome}")
    elif antecedents and outcomes:
        for antecedent in antecedents:
            for outcome in outcomes:
                signatures.append(f"{antecedent}->{outcome}")
    if boundaries and signatures:
        signatures.extend(f"{signature}|{boundary}" for signature in signatures[:2] for boundary in boundaries[:1])
    return signatures[:6]


def marker_sentences(text: str, markers: list[str], limit: int = 6) -> list[str]:
    sentences = extract_sentences(text, limit=260)
    hits: list[str] = []
    for sentence in sentences:
        lowered = sentence.lower()
        if any(marker.lower() in lowered for marker in markers):
            hits.append(sentence)
        if len(hits) >= limit:
            break
    return hits


def extract_hypotheses_propositions(text: str, matched: dict[str, list[str]], limit: int = 4) -> list[str]:
    markers = [
        "hypothesis",
        "hypotheses",
        "hypothesize",
        "proposition",
        "propositions",
        "we propose",
        "we argue",
        "命题",
        "命题一",
        "命题二",
        "假设",
        "h1",
        "h2",
    ]
    hits = marker_sentences(text, markers, limit=limit)
    if hits:
        return hits
    proposition = build_selective_proposition(matched)
    if proposition and "当前证据不足" not in proposition:
        return [proposition]
    return []


def extract_variable_roles(
    text: str,
    matched: dict[str, list[str]],
    hypothesis_sentences: list[str],
) -> dict[str, list[str]]:
    controls = marker_sentences(
        text,
        ["control", "controls", "covariate", "covariates", "控制变量", "控制了", "controlled for"],
        limit=4,
    )
    control_labels: list[str] = []
    for sentence in controls:
        lowered = sentence.lower()
        for category, labels in CODEBOOK.items():
            for label, keywords in labels.items():
                if any(keyword.lower() in lowered for keyword in keywords) and label not in control_labels:
                    control_labels.append(label)
    control_vars = control_labels[:4] or [truncate_text(sentence, 80) for sentence in controls[:3]]
    return {
        "independent_vars": matched.get("antecedents", [])[:6],
        "mediator_moderator_vars": (matched.get("mechanisms", []) + matched.get("boundaries", []))[:6],
        "dependent_vars": matched.get("outcomes", [])[:6],
        "control_vars": control_vars,
    }


def extract_future_research_items(text: str, matched: dict[str, list[str]], limit: int = 5) -> tuple[list[str], list[str]]:
    markers = [
        "future research",
        "future studies",
        "future study",
        "future work",
        "further research",
        "research agenda",
        "research direction",
        "limitation",
        "limitations",
        "should examine",
        "could explore",
        "未来研究",
        "未来可",
        "未来可以",
        "后续研究",
        "研究方向",
        "值得进一步",
        "有待进一步",
        "未来工作",
    ]
    future_sentences = marker_sentences(text, markers, limit=limit)
    if not future_sentences and matched.get("future_directions"):
        future_sentences = [f"未来研究可围绕 {'、'.join(matched['future_directions'][:3])} 展开。"]
    codes: list[str] = []
    for sentence in future_sentences:
        lowered = sentence.lower()
        for label, keywords in CODEBOOK.get("future_directions", {}).items():
            if any(keyword.lower() in lowered for keyword in keywords) and label not in codes:
                codes.append(label)
        for category in ("methods", "research_objects", "boundaries"):
            for label in matched.get(category, []):
                if label not in codes and len(codes) < 6:
                    codes.append(label)
    if not codes:
        codes = matched.get("future_directions", [])[:4]
    return future_sentences[:limit], codes[:6]


def build_axial_summary(matched: dict[str, list[str]]) -> str:
    chunks: list[str] = []
    labels = {
        "前因": matched.get("antecedents", []),
        "结果": matched.get("outcomes", []),
        "机制": matched.get("mechanisms", []),
        "边界条件": matched.get("boundaries", []),
        "研究对象": matched.get("research_objects", []),
        "方法": matched.get("methods", []),
        "理论": matched.get("theory_basis", []),
        "未来研究": matched.get("future_directions", []),
    }
    for name, items in labels.items():
        if items:
            chunks.append(f"{name}: {'、'.join(items[:3])}")
    return " | ".join(chunks) if chunks else "待人工复核"


def build_selective_summary(matched: dict[str, list[str]]) -> str:
    antecedent = "、".join(matched.get("antecedents", [])[:2])
    mechanism = "、".join(matched.get("mechanisms", [])[:2])
    outcome = "、".join(matched.get("outcomes", [])[:2])
    boundary = "、".join(matched.get("boundaries", [])[:2])
    if antecedent and outcome and mechanism and boundary:
        return f"该文主线可概括为：{antecedent} 通过 {mechanism} 影响 {outcome}，并受到 {boundary} 的约束。"
    if antecedent and outcome and mechanism:
        return f"该文主线可概括为：{antecedent} 通过 {mechanism} 影响 {outcome}。"
    if antecedent and outcome:
        return f"该文主线主要讨论 {antecedent} 与 {outcome} 之间的关系。"
    if outcome:
        return f"该文更强调 {outcome} 这一结果维度。"
    return "该文尚未形成稳定主轴概括，建议人工复核。"


def build_axial_relations(matched: dict[str, list[str]]) -> str:
    relation_parts: list[str] = []
    antecedent = "、".join(matched.get("antecedents", [])[:2])
    mechanism = "、".join(matched.get("mechanisms", [])[:2])
    outcome = "、".join(matched.get("outcomes", [])[:2])
    boundary = "、".join(matched.get("boundaries", [])[:2])
    if antecedent and mechanism and outcome:
        relation_parts.append(f"主链: {antecedent} -> {mechanism} -> {outcome}")
    elif antecedent and outcome:
        relation_parts.append(f"主链: {antecedent} -> {outcome}")
    if boundary:
        relation_parts.append(f"边界: {boundary}")
    if matched.get("research_objects"):
        relation_parts.append(f"对象: {'、'.join(matched['research_objects'][:2])}")
    if matched.get("methods"):
        relation_parts.append(f"方法: {'、'.join(matched['methods'][:2])}")
    if matched.get("theory_basis"):
        relation_parts.append(f"理论: {'、'.join(matched['theory_basis'][:2])}")
    return " | ".join(relation_parts) if relation_parts else "待人工复核"


def build_selective_proposition(matched: dict[str, list[str]]) -> str:
    antecedent = "、".join(matched.get("antecedents", [])[:2])
    mechanism = "、".join(matched.get("mechanisms", [])[:2])
    outcome = "、".join(matched.get("outcomes", [])[:2])
    boundary = "、".join(matched.get("boundaries", [])[:2])
    if antecedent and mechanism and outcome and boundary:
        return f"命题草案：在 {boundary} 情境下，{antecedent} 通过 {mechanism} 影响 {outcome}。"
    if antecedent and mechanism and outcome:
        return f"命题草案：{antecedent} 通过 {mechanism} 影响 {outcome}。"
    if antecedent and outcome:
        return f"命题草案：{antecedent} 与 {outcome} 之间存在值得进一步检验的关系。"
    return "命题草案：当前证据不足，建议人工复核后再抽象理论命题。"


def build_gap_focus(novel_themes: list[str], novel_relations: list[str], matched: dict[str, list[str]]) -> str:
    if novel_relations:
        return "变量关系链新增"
    if novel_themes:
        if matched.get("future_directions"):
            return "未来研究方向新增"
        if matched.get("mechanisms"):
            return "机制维度新增"
        if matched.get("boundaries"):
            return "边界条件新增"
        return "主题标签新增"
    return "支持性补充"


def build_coding_confidence(records: list[dict[str, str]], matched: dict[str, list[str]]) -> str:
    categories_hit = sum(1 for category in ("antecedents", "outcomes", "mechanisms", "boundaries") if matched.get(category))
    evidence_count = len(records)
    if categories_hit >= 3 and evidence_count >= 8:
        return "high"
    if categories_hit >= 2 and evidence_count >= 4:
        return "medium"
    return "low"


def build_recommendation(
    novel_themes: list[str],
    novel_relations: list[str],
    matched: dict[str, list[str]],
    baseline_ready: bool,
) -> tuple[str, str]:
    if novel_relations:
        relation_text = "；".join(novel_relations[:3])
        return "high", f"建议重点补充新的变量关系链 {relation_text}，可直接写入理论框架或研究命题部分。"
    if novel_themes:
        themes = "、".join(novel_themes[:4])
        if matched.get("future_directions"):
            return "high", f"建议重点补充 {themes}，并把它们写入你的研究不足或未来研究方向。"
        return "high", f"建议重点补充 {themes}，可作为你现有论文中尚未展开的新主题。"
    if not baseline_ready:
        return "medium", "当前未提供你的论文全文基线，已先按 covered_topics 粗略提醒；补充 baseline_paths 后提醒会更准。"
    return "low", "该文与当前已覆盖主题较接近，可优先作为支持性或补充性文献。"


def load_baseline_labels(config: dict[str, Any]) -> tuple[set[str], bool]:
    texts: list[str] = []
    ready = False
    for raw in config.get("baseline_paths", []):
        path = Path(raw)
        if not path.exists():
            continue
        text = read_text_file(path)
        if text.strip():
            texts.append(text)
            ready = True
    baseline_labels: set[str] = set()
    if texts:
        baseline_text = "\n".join(texts)
        baseline_labels = flatten_labels(match_labels(baseline_text))
    for topic in config.get("covered_topics", []):
        topic_hits = flatten_labels(match_labels(topic))
        if topic_hits:
            baseline_labels.update(topic_hits)
        else:
            baseline_labels.add(topic)
    return baseline_labels, ready


def load_baseline_relation_signatures(config: dict[str, Any]) -> set[str]:
    texts: list[str] = []
    for raw in config.get("baseline_paths", []):
        path = Path(raw)
        if not path.exists():
            continue
        text = read_text_file(path)
        if text.strip():
            texts.append(text)
    signatures: set[str] = set()
    if texts:
        baseline_text = "\n".join(texts)
        baseline_matched = match_labels(baseline_text)
        signatures.update(build_relation_signatures(baseline_matched))
    for topic in config.get("covered_topics", []):
        topic_matched = match_labels(topic)
        signatures.update(build_relation_signatures(topic_matched))
    return signatures


def default_skill_dirs() -> list[Path]:
    roots = [
        Path.home() / ".codex" / "skills",
        Path.cwd() / "skills",
    ]
    out: list[Path] = []
    seen: set[Path] = set()
    for root in roots:
        resolved = root.expanduser().resolve()
        if resolved in seen:
            continue
        seen.add(resolved)
        out.append(resolved)
    return out


def parse_skill_frontmatter(skill_md: Path) -> dict[str, Any]:
    content = skill_md.read_text(encoding="utf-8", errors="ignore")
    meta = {
        "name": skill_md.parent.name,
        "description": "",
        "path": str(skill_md.parent),
        "skill_md": str(skill_md),
    }
    if content.startswith("---"):
        parts = content.split("---", 2)
        if len(parts) >= 3:
            frontmatter = parts[1]
            for line in frontmatter.splitlines():
                if ":" not in line:
                    continue
                key, value = line.split(":", 1)
                key = key.strip()
                value = value.strip().strip('"').strip("'")
                if key in {"name", "description"} and value:
                    meta[key] = value
    if not meta["description"]:
        for line in content.splitlines():
            line = line.strip()
            if line and not line.startswith("#") and len(line) > 8:
                meta["description"] = line[:240]
                break
    return meta


def discover_skills(config: dict[str, Any], query_text: str) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    roots: list[Path] = []
    seen_roots: set[Path] = set()
    for raw in config.get("skill_dirs", []):
        resolved = Path(raw).expanduser().resolve()
        if resolved in seen_roots:
            continue
        seen_roots.add(resolved)
        roots.append(resolved)
    for fallback in default_skill_dirs():
        if fallback in seen_roots:
            continue
        seen_roots.add(fallback)
        roots.append(fallback)

    all_skills: list[dict[str, Any]] = []
    seen_paths: set[str] = set()
    q_tokens = set(query_tokens(query_text))
    for root in roots:
        if not root.exists():
            continue
        for skill_md in root.rglob("SKILL.md"):
            skill_path = str(skill_md.parent.resolve())
            if skill_path in seen_paths:
                continue
            seen_paths.add(skill_path)
            meta = parse_skill_frontmatter(skill_md)
            haystack = normalize_text(meta.get("name", "") + " " + meta.get("description", ""))
            score = 0
            if q_tokens:
                score = sum(1 for token in q_tokens if token in haystack)
            meta["score"] = score
            all_skills.append(meta)

    all_skills.sort(key=lambda item: (int(item.get("score", 0)), item.get("name", "")), reverse=True)
    selected = [item for item in all_skills if int(item.get("score", 0)) > 0][:8]
    if not selected:
        selected = all_skills[:5]
    return all_skills, selected


def guess_local_title(text: str, fallback: str) -> str:
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    for line in lines[:15]:
        if 8 <= len(line) <= 220 and not re.search(r"(abstract|摘要|关键词|doi|vol\.|issue)", line.lower()):
            return line
    return fallback


def guess_local_authors(text: str) -> str:
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    for line in lines[:12]:
        if 2 <= len(line) <= 120 and re.search(r"[A-Za-z\u4e00-\u9fff]", line) and "," in line:
            return line
    return ""


def translation_config(config: dict[str, Any]) -> dict[str, Any]:
    return config.get("translation", {}) or {}


def assistant_config(config: dict[str, Any]) -> dict[str, Any]:
    return config.get("assistant", {}) or {}


def call_openai_compatible_chat(
    settings: dict[str, Any],
    messages: list[dict[str, str]],
    *,
    temperature: float | None = None,
    max_tokens: int | None = None,
) -> str:
    api_key = os.environ.get(settings.get("api_key_env", "GROUNDED_AGENT_API_KEY"), "").strip()
    api_url = str(settings.get("api_url", "")).strip()
    model = str(settings.get("model", "")).strip()
    if not api_key or not api_url or not model:
        return ""

    payload: dict[str, Any] = {
        "model": model,
        "messages": messages,
        "temperature": float(settings.get("temperature", 0.2) if temperature is None else temperature),
    }
    if max_tokens:
        payload["max_tokens"] = int(max_tokens)

    req = urllib.request.Request(
        api_url.rstrip("/") + "/chat/completions",
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}",
        },
    )
    try:
        with urllib.request.urlopen(req, timeout=90) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        return data["choices"][0]["message"]["content"].strip()
    except Exception:
        return ""


def translate_text(
    text: str,
    config: dict[str, Any],
    cache: dict[str, str],
    cache_key_prefix: str,
) -> str:
    settings = translation_config(config)
    if not settings.get("enabled") or not text.strip():
        return ""
    digest = hashlib.md5(f"{cache_key_prefix}:{text}".encode("utf-8")).hexdigest()
    if digest in cache:
        return cache[digest]

    api_key = os.environ.get(settings.get("api_key_env", "GROUNDED_TRANSLATE_API_KEY"), "").strip()
    api_url = str(settings.get("api_url", "")).strip()
    model = str(settings.get("model", "")).strip()
    if not api_key or not api_url or not model:
        return ""

    translated = call_openai_compatible_chat(
        settings,
        [
            {
                "role": "user",
                "content": f"请将以下学术文本翻译成中文，只返回翻译结果，不要加解释：\n\n{text}",
            }
        ],
        temperature=0.0,
        max_tokens=2000,
    )
    if not translated:
        return ""

    cache[digest] = translated
    return translated


def load_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return default


def save_json(path: Path, data: Any) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def append_jsonl(path: Path, item: dict[str, Any]) -> None:
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(item, ensure_ascii=False) + "\n")


def update_state(path: Path, **kwargs: Any) -> dict[str, Any]:
    state = load_json(path, {})
    state.update(kwargs)
    save_json(path, state)
    return state


def load_translation_cache(path: Path) -> dict[str, str]:
    data = load_json(path, {})
    return data if isinstance(data, dict) else {}


def load_download_history(path: Path) -> dict[str, dict[str, Any]]:
    data = load_json(path, {})
    return data if isinstance(data, dict) else {}


def save_download_history(path: Path, data: dict[str, dict[str, Any]]) -> None:
    save_json(path, data)


def log_agent_trace(path: Path, step: str, status: str, **kwargs: Any) -> None:
    payload = {
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "step": step,
        "status": status,
    }
    payload.update(kwargs)
    append_jsonl(path, payload)


def build_agent_plan(config: dict[str, Any], selected_skills: list[dict[str, Any]]) -> list[dict[str, Any]]:
    plan: list[dict[str, Any]] = [
        {
            "step": "discover_skills",
            "goal": "扫描 skill 目录并挑选与当前主题相关的 skills",
            "skills": [item.get("name", "") for item in selected_skills],
        }
    ]
    for source in config.get("sources", []):
        plan.append(
            {
                "step": f"search_{source}",
                "goal": f"从 {source} 检索与主题相关的文献",
            }
        )
    plan.extend(
        [
            {"step": "grounded_coding", "goal": "对新增文献做开放编码、聚焦编码和主轴概括"},
            {"step": "gap_detection", "goal": "对比基线主题，输出新增主题和补写建议"},
            {"step": "memory_persist", "goal": "写入运行记忆、主题记忆和紧凑上下文摘要"},
        ]
    )
    return plan[: int(config.get("agent", {}).get("max_turns", 12))]


def get_url_json(url: str, headers: dict[str, str] | None = None, timeout: int = 45) -> dict[str, Any]:
    req = urllib.request.Request(url, headers=headers or {})
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return json.loads(resp.read().decode("utf-8"))


def search_openalex_works(query: str, from_date: str, to_date: str, per_page: int) -> list[SearchCandidate]:
    params = {
        "search": query,
        "per-page": str(per_page),
        "sort": "publication_date:desc",
        "filter": f"from_publication_date:{from_date},to_publication_date:{to_date},is_retracted:false,type:article",
        "mailto": "openalex@example.com",
    }
    url = f"https://api.openalex.org/works?{urllib.parse.urlencode(params)}"
    data = get_json(url)
    results: list[SearchCandidate] = []
    for work in data.get("results", []):
        primary = work.get("primary_location") or {}
        source = primary.get("source") or {}
        authors = []
        for item in work.get("authorships", [])[:8]:
            author = item.get("author") or {}
            if author.get("display_name"):
                authors.append(author["display_name"])
        primary_topic = (work.get("primary_topic") or {}).get("display_name") or ""
        topics = "; ".join(topic.get("display_name") for topic in work.get("topics", [])[:6] if topic.get("display_name"))
        results.append(
            SearchCandidate(
                query=query,
                source_name="OpenAlex",
                source_type="api",
                external_id=work.get("id") or "",
                doi=(work.get("doi") or "").replace("https://doi.org/", ""),
                title=work.get("display_name") or "",
                authors="; ".join(authors),
                journal=source.get("display_name") or "",
                year=str(work.get("publication_year") or ""),
                publication_date=work.get("publication_date") or "",
                language=work.get("language") or "",
                cited_by_count=int(work.get("cited_by_count") or 0),
                peer_reviewed="是",
                primary_topic=primary_topic,
                topics=topics,
                abstract=invert_abstract(work.get("abstract_inverted_index"))[:8000],
                source_url=work.get("id") or "",
                pdf_url=(work.get("best_oa_location") or {}).get("pdf_url") or "",
                local_pdf="",
                preview_text=invert_abstract(work.get("abstract_inverted_index")) or (work.get("display_name") or ""),
            )
        )
    return results


def search_arxiv_works(query: str, from_date: str, to_date: str, per_page: int) -> list[SearchCandidate]:
    start_dt = datetime.fromisoformat(from_date)
    end_dt = datetime.fromisoformat(to_date)
    params = {
        "search_query": f"all:{query}",
        "start": "0",
        "max_results": str(per_page),
        "sortBy": "submittedDate",
        "sortOrder": "descending",
    }
    url = "http://export.arxiv.org/api/query?" + urllib.parse.urlencode(params)
    req = urllib.request.Request(url, headers={"User-Agent": "grounded-monitor/1.0"})
    with urllib.request.urlopen(req, timeout=45) as resp:
        xml_text = resp.read().decode("utf-8")
    root = ET.fromstring(xml_text)
    ns = {"atom": "http://www.w3.org/2005/Atom"}
    results: list[SearchCandidate] = []
    for entry in root.findall("atom:entry", ns):
        entry_id = (entry.findtext("atom:id", default="", namespaces=ns) or "").strip()
        title = re.sub(r"\s+", " ", entry.findtext("atom:title", default="", namespaces=ns) or "").strip()
        summary = re.sub(r"\s+", " ", entry.findtext("atom:summary", default="", namespaces=ns) or "").strip()
        published = (entry.findtext("atom:published", default="", namespaces=ns) or "")[:10]
        if published:
            pub_dt = datetime.fromisoformat(published)
            if pub_dt < start_dt or pub_dt > end_dt:
                continue
        authors = "; ".join(
            author.findtext("atom:name", default="", namespaces=ns).strip()
            for author in entry.findall("atom:author", ns)
            if author.findtext("atom:name", default="", namespaces=ns).strip()
        )
        pdf_url = ""
        for link in entry.findall("atom:link", ns):
            if link.attrib.get("title") == "pdf":
                pdf_url = link.attrib.get("href", "")
        if not pdf_url and entry_id:
            pdf_url = entry_id.replace("/abs/", "/pdf/") + ".pdf"
        results.append(
            SearchCandidate(
                query=query,
                source_name="arXiv",
                source_type="api",
                external_id=entry_id,
                doi="",
                title=title,
                authors=authors,
                journal="arXiv",
                year=published[:4],
                publication_date=published,
                language="en",
                cited_by_count=0,
                peer_reviewed="否",
                primary_topic="",
                topics="",
                abstract=summary[:8000],
                source_url=entry_id,
                pdf_url=pdf_url,
                local_pdf="",
                preview_text=summary or title,
            )
        )
    return results


def search_semantic_scholar_works(query: str, from_date: str, to_date: str, per_page: int) -> list[SearchCandidate]:
    fields_param = ",".join(
        [
            "title",
            "abstract",
            "authors",
            "year",
            "venue",
            "publicationDate",
            "externalIds",
            "citationCount",
            "openAccessPdf",
            "url",
            "publicationTypes",
            "tldr",
        ]
    )
    params = {
        "query": query,
        "limit": str(per_page),
        "fields": fields_param,
        "year": f"{from_date[:4]}-{to_date[:4]}",
    }
    headers = {"User-Agent": "grounded-monitor/1.0"}
    api_key = os.environ.get("SEMANTIC_SCHOLAR_API_KEY", "").strip()
    if api_key:
        headers["x-api-key"] = api_key
    url = "https://api.semanticscholar.org/graph/v1/paper/search?" + urllib.parse.urlencode(params)
    data = get_url_json(url, headers=headers)
    results: list[SearchCandidate] = []
    for paper in data.get("data", []):
        publication_date = (paper.get("publicationDate") or "")[:10]
        if publication_date and not (from_date <= publication_date <= to_date):
            continue
        authors = "; ".join(author.get("name", "") for author in paper.get("authors", [])[:8] if author.get("name"))
        abstract = paper.get("abstract") or ((paper.get("tldr") or {}).get("text") or "")
        ext = paper.get("externalIds") or {}
        doi = (ext.get("DOI") or "").strip()
        pub_types = paper.get("publicationTypes") or []
        peer_reviewed = "是" if any(item in {"JournalArticle", "Conference"} for item in pub_types) else "待定"
        pdf_url = ((paper.get("openAccessPdf") or {}).get("url")) or ""
        results.append(
            SearchCandidate(
                query=query,
                source_name="Semantic Scholar",
                source_type="api",
                external_id=paper.get("paperId") or paper.get("url") or "",
                doi=doi,
                title=paper.get("title") or "",
                authors=authors,
                journal=paper.get("venue") or "",
                year=str(paper.get("year") or ""),
                publication_date=publication_date,
                language="en",
                cited_by_count=int(paper.get("citationCount") or 0),
                peer_reviewed=peer_reviewed,
                primary_topic="",
                topics="",
                abstract=abstract[:8000],
                source_url=paper.get("url") or "",
                pdf_url=pdf_url,
                local_pdf="",
                preview_text=abstract or (paper.get("title") or ""),
            )
        )
    return results


def local_library_roots(config: dict[str, Any]) -> list[Path]:
    roots: list[Path] = []
    for raw in config.get("local_library_paths", []):
        roots.append(Path(raw).expanduser().resolve())
    cwd = Path.cwd()
    for fallback in [cwd / "papers", cwd / "literature"]:
        if fallback not in roots:
            roots.append(fallback)
    return roots


def local_relevance_score(query: str, pdf_path: Path, preview_text: str) -> float:
    q_tokens = set(query_tokens(query))
    if not q_tokens:
        return 0.0
    haystack = normalize_text(pdf_path.stem + " " + preview_text)
    hits = sum(1 for token in q_tokens if token in haystack)
    return hits / max(len(q_tokens), 1)


def scan_local_library(config: dict[str, Any], query: str) -> list[SearchCandidate]:
    candidates: list[tuple[float, SearchCandidate]] = []
    seen_paths: set[Path] = set()
    for root in local_library_roots(config):
        if not root.exists():
            continue
        for pdf_path in root.rglob("*.pdf"):
            resolved = pdf_path.resolve()
            if resolved in seen_paths:
                continue
            seen_paths.add(resolved)
            try:
                preview_text = extract_pdf_text(resolved, page_limit=3)
            except Exception:
                continue
            score = local_relevance_score(query, resolved, preview_text)
            if score <= 0:
                continue
            title = guess_local_title(preview_text, resolved.stem)
            authors = guess_local_authors(preview_text)
            year_match = re.search(r"(20\d{2})", preview_text)
            year = year_match.group(1) if year_match else ""
            candidates.append(
                (
                    score,
                    SearchCandidate(
                        query=query,
                        source_name="Local Library",
                        source_type="local",
                        external_id=str(resolved),
                        doi="",
                        title=title,
                        authors=authors,
                        journal="",
                        year=year,
                        publication_date="",
                        language="",
                        cited_by_count=0,
                        peer_reviewed="待定",
                        primary_topic="",
                        topics="",
                        abstract="",
                        source_url=str(resolved),
                        pdf_url="",
                        local_pdf=str(resolved),
                        preview_text=preview_text[:12000],
                    ),
                )
            )
    candidates.sort(key=lambda item: item[0], reverse=True)
    return [item[1] for item in candidates[: int(config.get("max_local_papers", 20))]]


def fetch_candidates_for_source(
    source: str,
    query: str,
    config: dict[str, Any],
    from_date: str,
    to_date: str,
) -> list[SearchCandidate]:
    source = source.strip().lower()
    per_page = int(config["max_results_per_query"])
    if source == "openalex":
        return search_openalex_works(query, from_date, to_date, per_page)
    if source == "arxiv":
        return search_arxiv_works(query, from_date, to_date, per_page)
    if source in {"semantic_scholar", "semantic-scholar"}:
        return search_semantic_scholar_works(query, from_date, to_date, per_page)
    if source == "local":
        return scan_local_library(config, query)
    return []


def resolve_candidate_pdf(candidate: SearchCandidate) -> tuple[str, str]:
    if candidate.local_pdf:
        return candidate.local_pdf, "local_pdf"
    if candidate.pdf_url:
        return candidate.pdf_url, "direct_pdf"
    if candidate.source_name == "OpenAlex" and candidate.source_url:
        try:
            work = get_json(candidate.source_url)
        except Exception:
            return "", "openalex_fetch_failed"
        best_oa = work.get("best_oa_location") or {}
        pdf_url = best_oa.get("pdf_url") or ""
        if pdf_url:
            return pdf_url, "best_oa_location"
        landing = (
            best_oa.get("landing_page_url")
            or (work.get("primary_location") or {}).get("landing_page_url")
            or (f"https://doi.org/{candidate.doi}" if candidate.doi else "")
        )
        if not landing:
            return "", "no_landing_url"
        discovered, note = discover_pdf_url(landing)
        return discovered, note
    return "", "no_pdf_url"


def download_candidate_pdf(
    candidate: SearchCandidate,
    pdf_dir: Path,
    history: dict[str, dict[str, Any]],
) -> tuple[str, str, dict[str, dict[str, Any]]]:
    key = candidate_key(candidate.title, candidate.doi, candidate.external_id)
    if candidate.local_pdf:
        return candidate.local_pdf, "是", history
    if key in history:
        previous = history[key]
        file_path = previous.get("file_path", "")
        if file_path and Path(file_path).exists():
            return file_path, "是", history

    pdf_url, _note = resolve_candidate_pdf(candidate)
    if not pdf_url:
        return "", "否", history
    filename = f"{slugify(candidate.title, limit=120)}.pdf"
    dest = pdf_dir / filename
    ok, note = try_download(pdf_url, dest)
    if not ok:
        return "", "否", history
    history[key] = {
        "title": candidate.title,
        "file_path": str(dest),
        "pdf_url": pdf_url,
        "download_date": datetime.now().isoformat(timespec="seconds"),
        "source_name": candidate.source_name,
    }
    return str(dest), "是", history


def make_row(
    candidate: SearchCandidate,
    config: dict[str, Any],
    baseline_labels: set[str],
    baseline_relation_signatures: set[str],
    baseline_ready: bool,
    pdf_dir: Path,
    download_history: dict[str, dict[str, Any]],
    translation_cache: dict[str, str],
) -> tuple[MonitorRow, dict[str, dict[str, Any]], dict[str, str]]:
    local_pdf = candidate.local_pdf
    pdf_downloaded = "否"
    full_text = candidate.preview_text or candidate.abstract or candidate.title

    if config.get("download_pdfs"):
        local_pdf, pdf_downloaded, download_history = download_candidate_pdf(candidate, pdf_dir, download_history)
        if local_pdf and Path(local_pdf).exists():
            try:
                full_text = extract_pdf_text(Path(local_pdf), page_limit=int(config.get("max_pdf_pages_for_coding", 30)))
            except Exception:
                full_text = candidate.preview_text or candidate.abstract or candidate.title
    elif local_pdf and Path(local_pdf).exists():
        try:
            full_text = extract_pdf_text(Path(local_pdf), page_limit=int(config.get("max_pdf_pages_for_coding", 30)))
        except Exception:
            full_text = candidate.preview_text or candidate.title

    coding_text = full_text or candidate.abstract or candidate.title
    records = build_open_code_records(coding_text)
    matched = aggregate_records(records)
    if not any(matched.values()):
        matched = match_labels(coding_text)
    open_codes = summarize_open_codes(matched)
    evidence = best_evidence(coding_text, matched)
    matched_labels = flatten_labels(matched)
    novel_themes = sorted(label for label in matched_labels if label not in baseline_labels)
    relation_signatures = build_relation_signatures(matched)
    novel_relations = sorted(signature for signature in relation_signatures if signature not in baseline_relation_signatures)
    alert_level, recommendation = build_recommendation(novel_themes, novel_relations, matched, baseline_ready)
    hypotheses_propositions = extract_hypotheses_propositions(coding_text, matched)
    variable_roles = extract_variable_roles(coding_text, matched, hypotheses_propositions)
    future_research_items, future_direction_codes = extract_future_research_items(coding_text, matched)

    title_zh = translate_text(candidate.title, config, translation_cache, "title")
    abstract_zh = translate_text(candidate.abstract[:2000], config, translation_cache, "abstract")

    row = MonitorRow(
        run_date=date.today().isoformat(),
        query=candidate.query,
        source_name=candidate.source_name,
        source_type=candidate.source_type,
        project_name=config["project_name"],
        external_id=candidate.external_id,
        doi=candidate.doi,
        title=candidate.title,
        title_zh=title_zh,
        authors=candidate.authors,
        journal=candidate.journal,
        year=candidate.year,
        publication_date=candidate.publication_date,
        language=candidate.language,
        cited_by_count=int(candidate.cited_by_count or 0),
        peer_reviewed=candidate.peer_reviewed,
        primary_topic=candidate.primary_topic,
        topics=candidate.topics,
        abstract=candidate.abstract[:4000],
        abstract_zh=abstract_zh[:4000],
        source_url=candidate.source_url,
        pdf_url=candidate.pdf_url,
        pdf_downloaded=pdf_downloaded,
        local_pdf=local_pdf,
        open_codes=safe_join(open_codes),
        axial_summary=build_axial_summary(matched),
        selective_summary=build_selective_summary(matched),
        antecedents=safe_join(matched.get("antecedents", [])),
        outcomes=safe_join(matched.get("outcomes", [])),
        mechanisms=safe_join(matched.get("mechanisms", [])),
        boundaries=safe_join(matched.get("boundaries", [])),
        future_directions=safe_join(matched.get("future_directions", [])),
        research_objects=safe_join(matched.get("research_objects", [])),
        methods=safe_join(matched.get("methods", [])),
        theory_basis=safe_join(matched.get("theory_basis", [])),
        novel_themes=safe_join(novel_themes),
        novel_relations=safe_join(novel_relations),
        gap_focus=build_gap_focus(novel_themes, novel_relations, matched),
        alert_level=alert_level,
        recommendation=recommendation,
        hypotheses_propositions=" || ".join(hypotheses_propositions[:4]),
        independent_vars=safe_join(variable_roles.get("independent_vars", [])),
        mediator_moderator_vars=safe_join(variable_roles.get("mediator_moderator_vars", [])),
        dependent_vars=safe_join(variable_roles.get("dependent_vars", [])),
        control_vars=safe_join(variable_roles.get("control_vars", [])),
        future_research_directions=" || ".join(future_research_items[:5]),
        future_direction_codes=safe_join(future_direction_codes),
        open_code_details=build_open_code_details(records),
        axial_relations=build_axial_relations(matched),
        selective_proposition=build_selective_proposition(matched),
        coding_confidence=build_coding_confidence(records, matched),
        evidence_sentences=" | ".join(evidence),
    )
    return row, download_history, translation_cache


def load_existing_rows(path: Path) -> list[MonitorRow]:
    if not path.exists():
        return []
    allowed = {field.name for field in fields(MonitorRow)}
    rows: list[MonitorRow] = []
    with path.open(encoding="utf-8-sig", newline="") as f:
        for raw in csv.DictReader(f):
            data = {name: raw.get(name, "") for name in allowed}
            try:
                data["cited_by_count"] = int(data.get("cited_by_count") or 0)
            except Exception:
                data["cited_by_count"] = 0
            rows.append(MonitorRow(**data))
    return rows


def write_csv(rows: list[MonitorRow], path: Path) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=SOURCE_COLUMNS)
        writer.writeheader()
        for row in rows:
            writer.writerow(asdict(row))


def write_xlsx(rows: list[MonitorRow], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "literature_table"
    ws.append(SOURCE_COLUMNS)
    for row in rows:
        ws.append([getattr(row, col, "") for col in SOURCE_COLUMNS])

    alerts = wb.create_sheet("gap_alerts")
    alert_cols = ["run_date", "title", "source_name", "novel_themes", "alert_level", "recommendation", "query", "doi"]
    alerts.append(alert_cols)
    for row in rows:
        if row.alert_level == "high":
            alerts.append([getattr(row, col, "") for col in alert_cols])

    summary_sheet = wb.create_sheet("source_summary")
    summary_sheet.append(["source_name", "count"])
    for source_name, count in Counter(row.source_name for row in rows).most_common():
        summary_sheet.append([source_name, count])

    coding_sheet = wb.create_sheet("open_coding")
    coding_sheet.append(["title", "category", "label", "evidence", "coding_confidence", "gap_focus"])
    for row in rows:
        details = [part.strip() for part in (row.open_code_details or "").split("||") if part.strip()]
        if not details:
            coding_sheet.append([row.title, "", "", "", row.coding_confidence, row.gap_focus])
            continue
        for detail in details:
            parts = [part.strip() for part in detail.split("|", 2)]
            category = parts[0] if len(parts) > 0 else ""
            label = parts[1] if len(parts) > 1 else ""
            evidence = parts[2] if len(parts) > 2 else ""
            coding_sheet.append([row.title, category, label, evidence, row.coding_confidence, row.gap_focus])

    axial_sheet = wb.create_sheet("axial_matrix")
    axial_cols = [
        "title",
        "antecedents",
        "mechanisms",
        "outcomes",
        "boundaries",
        "research_objects",
        "methods",
        "theory_basis",
        "axial_relations",
        "selective_proposition",
        "novel_relations",
    ]
    axial_sheet.append(axial_cols)
    for row in rows:
        axial_sheet.append([getattr(row, col, "") for col in axial_cols])

    logic_sheet = wb.create_sheet("research_logic")
    logic_cols = [
        "title",
        "hypotheses_propositions",
        "independent_vars",
        "mediator_moderator_vars",
        "dependent_vars",
        "control_vars",
        "future_research_directions",
        "future_direction_codes",
        "gap_focus",
        "recommendation",
    ]
    logic_sheet.append(logic_cols)
    for row in rows:
        logic_sheet.append([getattr(row, col, "") for col in logic_cols])

    for sheet in wb.worksheets:
        for column_cells in sheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 45)
            sheet.column_dimensions[column_cells[0].column_letter].width = width
    wb.save(path)


def write_snapshot(
    config: dict[str, Any],
    baseline_labels: set[str],
    baseline_relation_signatures: set[str],
    outdir: Path,
    discovered_skills: list[dict[str, Any]],
    selected_skills: list[dict[str, Any]],
    agent_plan: list[dict[str, Any]],
) -> None:
    snapshot = {
        "config": config,
        "builtin_skills": BUILTIN_SKILLS,
        "prompt_templates": PROMPT_TEMPLATES,
        "baseline_labels": sorted(baseline_labels),
        "baseline_relation_signatures": sorted(baseline_relation_signatures),
        "discovered_skills": discovered_skills,
        "selected_skills": selected_skills,
        "agent_plan": agent_plan,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
    }
    save_json(outdir / "builtin_skills_and_prompts.json", snapshot)


def write_theme_memory(rows: list[MonitorRow], path: Path) -> dict[str, Any]:
    memory = {
        "antecedents": Counter(),
        "outcomes": Counter(),
        "mechanisms": Counter(),
        "boundaries": Counter(),
        "future_directions": Counter(),
        "research_objects": Counter(),
        "methods": Counter(),
        "theory_basis": Counter(),
    }
    for row in rows:
        for key in memory:
            raw = getattr(row, key, "") or ""
            for item in [part.strip() for part in raw.split("；") if part.strip()]:
                memory[key][item] += 1
    serializable = {
        key: dict(counter.most_common())
        for key, counter in memory.items()
    }
    save_json(path, serializable)
    return serializable


def write_agent_memory(
    path: Path,
    config: dict[str, Any],
    rows: list[MonitorRow],
    new_rows: list[MonitorRow],
    theme_memory: dict[str, Any],
    selected_skills: list[dict[str, Any]],
) -> dict[str, Any]:
    previous = load_json(path, {})
    previous_runs = previous.get("recent_runs", [])
    summary = {
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "project_name": config["project_name"],
        "queries": config["queries"],
        "sources": config["sources"],
        "new_rows": len(new_rows),
        "total_rows": len(rows),
        "top_alert_titles": [row.title for row in new_rows if row.alert_level == "high"][:5],
        "selected_skills": [item.get("name", "") for item in selected_skills],
    }
    recent_runs = (previous_runs + [summary])[-10:]
    memory = {
        "project_name": config["project_name"],
        "last_updated": datetime.now().isoformat(timespec="seconds"),
        "recent_runs": recent_runs,
        "persistent_top_themes": theme_memory,
        "selected_skills": [item.get("name", "") for item in selected_skills],
        "notes": "该文件是轻量长期记忆，保留最近运行摘要、技能选择和高频主题，供后续运行快速恢复上下文。",
    }
    save_json(path, memory)
    return memory


def write_compact_context(
    path: Path,
    config: dict[str, Any],
    rows: list[MonitorRow],
    new_rows: list[MonitorRow],
    theme_memory: dict[str, Any],
    agent_memory: dict[str, Any],
    selected_skills: list[dict[str, Any]],
) -> str:
    settings = config.get("context_compression", {}) or {}
    enabled = bool(settings.get("enabled", True))
    char_threshold = int(settings.get("char_threshold", 30000))
    max_recent_rows = int(settings.get("max_recent_rows", 12))
    raw_size = sum(
        len(row.title or "") + len(row.abstract or "") + len(row.evidence_sentences or "")
        for row in rows
    )
    lines = [
        f"# Compact Context - {config['project_name']}",
        "",
        f"- generated_at: {datetime.now().isoformat(timespec='seconds')}",
        f"- compression_enabled: {'yes' if enabled else 'no'}",
        f"- raw_context_estimate_chars: {raw_size}",
        f"- threshold: {char_threshold}",
        f"- queries: {'；'.join(config['queries'])}",
        f"- sources: {'；'.join(config['sources'])}",
        f"- selected_skills: {'、'.join(item.get('name', '') for item in selected_skills) or 'none'}",
        "",
        "## Goal",
        "每天增量检索相关文献，进行扎根式编码，识别新增主题，并输出长期可恢复的研究记忆。",
        "",
        "## Top Themes",
    ]
    for category, mapping in theme_memory.items():
        top_items = list(mapping.items())[:5]
        if top_items:
            lines.append(f"- {category}: " + "；".join(f"{name}({count})" for name, count in top_items))
    lines.extend(["", "## Recent High Alerts"])
    alerts = [row for row in new_rows if row.alert_level == "high"][:8]
    if not alerts:
        lines.append("- none")
    for row in alerts:
        lines.append(f"- {row.title}: {row.novel_themes} -> {row.recommendation}")
    lines.extend(["", "## Recent Rows"])
    for row in rows[:max_recent_rows]:
        lines.append(
            f"- [{row.source_name}] {row.title} | {row.axial_relations or row.axial_summary} | "
            f"{row.novel_relations or row.novel_themes or '无新增主题'}"
        )
    lines.extend(["", "## Memory Summary"])
    for item in agent_memory.get("recent_runs", [])[-3:]:
        lines.append(
            f"- {item.get('timestamp')}: new={item.get('new_rows')} total={item.get('total_rows')} skills={','.join(item.get('selected_skills', []))}"
        )
    if enabled and raw_size >= char_threshold:
        lines.insert(0, "> Context compression triggered because the estimated context exceeded the configured threshold.\n")
    path.write_text("\n".join(lines).strip() + "\n", encoding="utf-8")
    return str(path)


def row_search_text(row: MonitorRow) -> str:
    return normalize_text(
        " ".join(
            [
                row.title,
                row.title_zh,
                row.query,
                row.primary_topic,
                row.topics,
                row.abstract,
                row.abstract_zh,
                row.axial_summary,
                row.selective_summary,
                row.antecedents,
                row.outcomes,
                row.mechanisms,
                row.boundaries,
                row.future_directions,
                row.research_objects,
                row.methods,
                row.theory_basis,
                row.novel_themes,
                row.novel_relations,
                row.gap_focus,
                row.recommendation,
                row.hypotheses_propositions,
                row.independent_vars,
                row.mediator_moderator_vars,
                row.dependent_vars,
                row.control_vars,
                row.future_research_directions,
                row.future_direction_codes,
                row.open_code_details,
                row.axial_relations,
                row.selective_proposition,
                row.coding_confidence,
                row.evidence_sentences,
            ]
        )
    )


def score_row_for_query(row: MonitorRow, query: str) -> float:
    tokens = set(query_tokens(query))
    if not tokens:
        return 0.0
    haystack = row_search_text(row)
    score = 0.0
    for token in tokens:
        if token in haystack:
            score += 1.0
    if row.alert_level == "high":
        score += 0.5
    if row.peer_reviewed == "是":
        score += 0.3
    return score


def row_reference(row: MonitorRow, index: int) -> str:
    year = row.year or row.publication_date[:4] or "n.d."
    source = row.journal or row.source_name or "unknown source"
    return f"[{index}] {row.title} ({year}, {source})"


def extract_pdf_excerpt_for_context(pdf_path: str, char_limit: int = 2200) -> str:
    if not pdf_path:
        return ""
    path = Path(pdf_path)
    if not path.exists():
        return ""
    try:
        text = extract_pdf_text(path, page_limit=4)
    except Exception:
        return ""
    text = re.sub(r"\s+", " ", text).strip()
    return text[:char_limit]


def build_row_context(row: MonitorRow, index: int) -> str:
    excerpt = extract_pdf_excerpt_for_context(row.local_pdf)
    fields = [
        row_reference(row, index),
        f"query={row.query}",
        f"source={row.source_name}",
        f"topic={row.primary_topic or row.topics or '待补充'}",
        f"axial={row.axial_summary or '待补充'}",
        f"selective={row.selective_summary or '待补充'}",
        f"antecedents={row.antecedents or '无'}",
        f"outcomes={row.outcomes or '无'}",
        f"mechanisms={row.mechanisms or '无'}",
        f"boundaries={row.boundaries or '无'}",
        f"future={row.future_directions or '无'}",
        f"research_objects={row.research_objects or '无'}",
        f"methods={row.methods or '无'}",
        f"theory_basis={row.theory_basis or '无'}",
        f"novel_themes={row.novel_themes or '无'}",
        f"novel_relations={row.novel_relations or '无'}",
        f"gap_focus={row.gap_focus or '无'}",
        f"recommendation={row.recommendation or '无'}",
        f"hypotheses_propositions={row.hypotheses_propositions or '无'}",
        f"independent_vars={row.independent_vars or '无'}",
        f"mediator_moderator_vars={row.mediator_moderator_vars or '无'}",
        f"dependent_vars={row.dependent_vars or '无'}",
        f"control_vars={row.control_vars or '无'}",
        f"future_research_directions={row.future_research_directions or '无'}",
        f"future_direction_codes={row.future_direction_codes or '无'}",
        f"axial_relations={row.axial_relations or '无'}",
        f"selective_proposition={row.selective_proposition or '无'}",
        f"coding_confidence={row.coding_confidence or '无'}",
        f"evidence={row.evidence_sentences or '无'}",
        f"abstract={(row.abstract_zh or row.abstract or '无')[:1800]}",
    ]
    if excerpt:
        fields.append(f"pdf_excerpt={excerpt}")
    return "\n".join(fields)


def select_relevant_rows(rows: list[MonitorRow], query: str, limit: int) -> list[MonitorRow]:
    ranked = sorted(rows, key=lambda row: (score_row_for_query(row, query), row.cited_by_count), reverse=True)
    return [row for row in ranked if score_row_for_query(row, query) > 0][:limit] or ranked[:limit]


def build_context_bundle(
    rows: list[MonitorRow],
    query: str,
    *,
    max_rows: int,
    max_chars: int,
) -> tuple[str, list[MonitorRow]]:
    selected = select_relevant_rows(rows, query, max_rows)
    blocks: list[str] = []
    total = 0
    trimmed: list[MonitorRow] = []
    for idx, row in enumerate(selected, start=1):
        block = build_row_context(row, idx)
        if blocks and total + len(block) > max_chars:
            break
        blocks.append(block)
        total += len(block)
        trimmed.append(row)
    return "\n\n".join(blocks), trimmed


def fallback_answer_markdown(question: str, selected_rows: list[MonitorRow]) -> str:
    if not selected_rows:
        return (
            "## 直接回答\n"
            "当前文献库里没有检索到足够相关的证据，暂时不能可靠回答这个问题。\n\n"
            "## 建议\n"
            "- 扩展检索词后再运行一次监测\n"
            "- 补充本地 PDF、Zotero 或更贴近问题的行业资料\n"
        )

    antecedents = Counter()
    outcomes = Counter()
    mechanisms = Counter()
    boundaries = Counter()
    for row in selected_rows:
        for raw, counter in [
            (row.antecedents, antecedents),
            (row.outcomes, outcomes),
            (row.mechanisms, mechanisms),
            (row.boundaries, boundaries),
        ]:
            for item in [part.strip() for part in raw.split("；") if part.strip()]:
                counter[item] += 1

    direct = []
    if antecedents:
        direct.append("高频驱动因素集中在 " + "、".join(name for name, _ in antecedents.most_common(3)))
    if mechanisms:
        direct.append("主要作用机制包括 " + "、".join(name for name, _ in mechanisms.most_common(3)))
    if outcomes:
        direct.append("常见结果变量包括 " + "、".join(name for name, _ in outcomes.most_common(3)))
    if boundaries:
        direct.append("边界条件多出现在 " + "、".join(name for name, _ in boundaries.most_common(3)))

    lines = [
        "## 直接回答",
        "；".join(direct) if direct else "现有证据更适合做方向性判断，仍需要人工复核原文。",
        "",
        "## 可转写命题",
    ]
    propositions = [row.selective_proposition for row in selected_rows if row.selective_proposition][:4]
    for proposition in propositions:
        lines.append(f"- {proposition}")
    if not propositions:
        lines.append("- 当前尚未形成足够稳定的命题草案。")

    lines.extend(["", "## 变量提取"])
    for row in selected_rows[:4]:
        lines.append(f"- {row.title}")
        lines.append(f"  自变量：{row.independent_vars or '待补充'}")
        lines.append(f"  中介/调节：{row.mediator_moderator_vars or '待补充'}")
        lines.append(f"  因变量：{row.dependent_vars or '待补充'}")
        lines.append(f"  控制变量：{row.control_vars or '待补充'}")

    lines.extend(["", "## 未来研究方向初级编码"])
    future_lines = False
    for row in selected_rows[:4]:
        if not row.future_research_directions and not row.future_direction_codes:
            continue
        future_lines = True
        lines.append(f"- {row.title}")
        lines.append(f"  未来研究：{row.future_research_directions or '待补充'}")
        lines.append(f"  初级编码：{row.future_direction_codes or '待补充'}")
    if not future_lines:
        lines.append("- 当前未提取到足够明确的未来研究方向。")

    lines.extend(["", "## 关键证据"])
    for idx, row in enumerate(selected_rows[:6], start=1):
        lines.append(f"- {row_reference(row, idx)}")
        lines.append(f"  结论：{row.selective_summary or row.axial_summary or '待人工复核'}")
        lines.append(f"  关系链：{row.axial_relations or '待补充'}")
        lines.append(f"  证据：{row.evidence_sentences or row.abstract[:180] or '待补充'}")

    lines.extend(["", "## 局限与下一步"])
    lines.append("- 当前回答优先基于文献摘要、编码结果和少量 PDF 片段，不等于完整全文审读。")
    lines.append("- 如果你希望把回答直接写进论文，可以再要求我输出“理论命题版”或“文献综述版”。")
    return "\n".join(lines)


def fallback_industry_report_markdown(
    topic: str,
    selected_rows: list[MonitorRow],
    theme_memory: dict[str, Any],
) -> str:
    lines = [
        f"# 行业报告：{topic}",
        "",
        "## 执行摘要",
    ]
    if not selected_rows:
        lines.append("当前文献库中与该主题直接相关的材料较少，建议先扩展检索词与数据源。")
        return "\n".join(lines) + "\n"

    top_titles = "；".join(row.title for row in selected_rows[:3])
    lines.append(f"本报告基于当前文献库中最相关的 {len(selected_rows)} 篇材料形成初版判断，核心参考包括：{top_titles}。")

    lines.extend(["", "## 行业现状"])
    for row in selected_rows[:4]:
        lines.append(f"- {row.title}：{row.selective_summary or row.axial_summary or row.recommendation}")

    lines.extend(["", "## 关键驱动因素"])
    for category in ["antecedents", "mechanisms", "outcomes", "boundaries", "research_objects", "methods", "theory_basis"]:
        mapping = theme_memory.get(category) or {}
        top_items = list(mapping.items())[:5]
        if not top_items:
            continue
        label = {
            "antecedents": "驱动因素",
            "mechanisms": "作用机制",
            "outcomes": "结果表现",
            "boundaries": "边界条件",
            "research_objects": "研究对象",
            "methods": "研究方法",
            "theory_basis": "理论基础",
        }[category]
        lines.append(f"- {label}：" + "；".join(f"{name}({count})" for name, count in top_items))

    lines.extend(["", "## 风险与不确定性"])
    risks = [row for row in selected_rows if row.boundaries or row.future_directions]
    if not risks:
        lines.append("- 当前证据对风险的直接讨论有限，建议补充政策、市场和竞争对手数据。")
    for row in risks[:4]:
        risk_items: list[str] = []
        for raw in [row.boundaries, row.future_directions, row.novel_themes]:
            for item in [part.strip() for part in raw.split("；") if part.strip()]:
                if item not in risk_items:
                    risk_items.append(item)
        risk_text = "；".join(risk_items)
        lines.append(f"- {row.title}：{risk_text or '待补充'}")

    lines.extend(["", "## 建议"])
    lines.append("- 把高频前因与机制转成行业观察框架，持续跟踪是否出现新的变量关系。")
    lines.append("- 对于高 alert 文献，优先检查其 novel_relations 和 selective_proposition，直接转写进论文的理论命题或综述表。")
    lines.append("- 若用于正式行业报告，建议补充新闻、财报、政策文件和访谈材料。")
    return "\n".join(lines) + "\n"


def answer_question(
    question: str,
    rows: list[MonitorRow],
    config: dict[str, Any],
    outdir: Path,
    agent_trace_path: Path,
) -> str:
    settings = assistant_config(config)
    context, selected_rows = build_context_bundle(
        rows,
        question,
        max_rows=int(settings.get("max_context_rows", 10)),
        max_chars=int(settings.get("max_context_chars", 24000)),
    )
    log_agent_trace(
        agent_trace_path,
        step="qa_answer",
        status="start",
        question=question,
        selected_rows=len(selected_rows),
    )

    answer_body = ""
    if settings.get("enabled"):
        answer_body = call_openai_compatible_chat(
            settings,
            [
                {"role": "system", "content": PROMPT_TEMPLATES["qa_answer"]},
                {
                    "role": "user",
                    "content": (
                        f"用户问题：{question}\n\n"
                        "请根据以下证据回答，并在关键判断后用 [1] [2] 这样的编号标注来源。"
                        "若证据不足，请明确写出“证据不足”。\n\n"
                        f"{context}"
                    ),
                },
            ],
            temperature=float(settings.get("temperature", 0.2)),
            max_tokens=3200,
        )
    if not answer_body:
        answer_body = fallback_answer_markdown(question, selected_rows)

    references = [f"- {row_reference(row, idx)}" for idx, row in enumerate(selected_rows, start=1)]
    lines = [
        f"# 问答结果：{question}",
        "",
        f"- generated_at: {datetime.now().isoformat(timespec='seconds')}",
        f"- selected_sources: {len(selected_rows)}",
        "",
        answer_body.strip(),
        "",
        "## 参考来源",
    ]
    lines.extend(references or ["- 当前没有可引用来源。"])

    answer_dir = outdir / str(settings.get("answer_dirname", "qa_answers"))
    answer_dir.mkdir(parents=True, exist_ok=True)
    filename = f"qa_{slugify(question, limit=80)}_{date.today().isoformat()}.md"
    output_path = answer_dir / filename
    output_path.write_text("\n".join(lines).strip() + "\n", encoding="utf-8")
    log_agent_trace(
        agent_trace_path,
        step="qa_answer",
        status="ok",
        question=question,
        output_path=str(output_path),
    )
    return str(output_path)


def generate_industry_report(
    topic: str,
    rows: list[MonitorRow],
    config: dict[str, Any],
    outdir: Path,
    theme_memory: dict[str, Any],
    agent_trace_path: Path,
) -> str:
    settings = assistant_config(config)
    context, selected_rows = build_context_bundle(
        rows,
        topic,
        max_rows=max(int(settings.get("max_context_rows", 10)), 12),
        max_chars=int(settings.get("max_context_chars", 24000)),
    )
    log_agent_trace(
        agent_trace_path,
        step="industry_report",
        status="start",
        topic=topic,
        selected_rows=len(selected_rows),
    )

    report_body = ""
    if settings.get("enabled"):
        report_body = call_openai_compatible_chat(
            settings,
            [
                {"role": "system", "content": PROMPT_TEMPLATES["industry_report"]},
                {
                    "role": "user",
                    "content": (
                        f"报告主题：{topic}\n\n"
                        "请写成结构化 Markdown 报告，至少包含：执行摘要、行业现状、关键驱动因素、竞争格局、风险与不确定性、"
                        "机会与建议、后续跟踪指标。所有关键结论优先引用给定文献证据，不要编造财务数据。\n\n"
                        f"{context}"
                    ),
                },
            ],
            temperature=float(settings.get("temperature", 0.2)),
            max_tokens=4200,
        )
    if not report_body:
        report_body = fallback_industry_report_markdown(topic, selected_rows, theme_memory)

    references = [f"- {row_reference(row, idx)}" for idx, row in enumerate(selected_rows, start=1)]
    lines = [report_body.strip(), "", "## 参考来源"]
    lines.extend(references or ["- 当前没有可引用来源。"])

    report_dir = outdir / str(settings.get("report_dirname", "industry_reports"))
    report_dir.mkdir(parents=True, exist_ok=True)
    filename = f"industry_report_{slugify(topic, limit=80)}_{date.today().isoformat()}.md"
    output_path = report_dir / filename
    output_path.write_text("\n".join(lines).strip() + "\n", encoding="utf-8")
    log_agent_trace(
        agent_trace_path,
        step="industry_report",
        status="ok",
        topic=topic,
        output_path=str(output_path),
    )
    return str(output_path)


def write_daily_report(
    rows: list[MonitorRow],
    new_rows: list[MonitorRow],
    baseline_labels: set[str],
    baseline_ready: bool,
    report_path: Path,
    config: dict[str, Any],
    theme_memory: dict[str, Any],
    resumed: bool,
    selected_skills: list[dict[str, Any]],
    compact_context_path: str,
    agent_memory_path: str,
) -> None:
    query_counter = Counter(row.query for row in new_rows)
    source_counter = Counter(row.source_name for row in new_rows)
    lines = [
        f"# {config['project_name']} - {date.today().isoformat()}",
        "",
        "## 内置技能链",
    ]
    for item in BUILTIN_SKILLS:
        lines.append(f"- `{item['name']}`: {item['goal']}")

    lines.extend(["", "## 选中的外部 Skills"])
    if selected_skills:
        for item in selected_skills[:8]:
            lines.append(f"- `{item.get('name', '')}`: {item.get('description', '')}")
    else:
        lines.append("- 本次未发现额外外部 skills，使用内置能力完成。")

    lines.extend(
        [
            "",
            "## 本次运行概况",
            f"- 新增文献数：{len(new_rows)}",
            f"- 累积文献数：{len(rows)}",
            f"- 查询主题：{'；'.join(config['queries'])}",
            f"- 检索来源：{'；'.join(config['sources'])}",
            f"- 对比基线就绪：{'是' if baseline_ready else '否'}",
            f"- 是否从上次中断状态恢复：{'是' if resumed else '否'}",
            f"- 已覆盖主题标签：{'、'.join(sorted(baseline_labels)) if baseline_labels else '暂未提供'}",
            f"- 紧凑上下文文件：{compact_context_path}",
            f"- 长期记忆文件：{agent_memory_path}",
            "",
            "## 各查询新增数量",
        ]
    )
    for query, value in query_counter.items():
        lines.append(f"- {query}: {value}")

    lines.extend(["", "## 各来源新增数量"])
    for source_name, value in source_counter.items():
        lines.append(f"- {source_name}: {value}")

    lines.extend(["", "## 新主题提醒"])
    flagged = [row for row in new_rows if row.alert_level == "high"]
    if not flagged:
        lines.append("- 本次没有识别到明显超出既有主题边界的新主题。")
    for row in flagged[:10]:
        lines.append(f"- {row.title}")
        lines.append(f"  缺口类型：{row.gap_focus or '待补充'}")
        lines.append(f"  新主题：{row.novel_themes}")
        if row.novel_relations:
            lines.append(f"  新关系：{row.novel_relations}")
        lines.append(f"  建议：{row.recommendation}")

    lines.extend(["", "## 假设/命题与变量提取"])
    for row in new_rows[:8]:
        lines.append(f"- {row.title}")
        lines.append(f"  假设/命题：{row.hypotheses_propositions or row.selective_proposition or '待补充'}")
        lines.append(f"  自变量：{row.independent_vars or '待补充'}")
        lines.append(f"  中介/调节：{row.mediator_moderator_vars or '待补充'}")
        lines.append(f"  因变量：{row.dependent_vars or '待补充'}")
        if row.control_vars:
            lines.append(f"  控制变量：{row.control_vars}")

    lines.extend(["", "## 未来研究方向初级编码"])
    future_rows = [row for row in new_rows if row.future_research_directions or row.future_direction_codes][:10]
    if not future_rows:
        lines.append("- 本次没有提取到足够明确的未来研究方向。")
    for row in future_rows:
        lines.append(f"- {row.title}")
        lines.append(f"  未来研究：{row.future_research_directions or '待补充'}")
        lines.append(f"  初级编码：{row.future_direction_codes or '待补充'}")

    lines.extend(["", "## 主题记忆"])
    for category, mapping in theme_memory.items():
        top_items = list(mapping.items())[:5]
        if not top_items:
            continue
        label = {
            "antecedents": "前因",
            "outcomes": "结果",
            "mechanisms": "机制",
            "boundaries": "边界条件",
            "future_directions": "未来研究",
            "research_objects": "研究对象",
            "methods": "研究方法",
            "theory_basis": "理论基础",
        }[category]
        lines.append(f"- {label}: " + "；".join(f"{name}({count})" for name, count in top_items))

    lines.extend(["", "## 新增关系链"])
    relation_rows = [row for row in new_rows if row.novel_relations][:10]
    if not relation_rows:
        lines.append("- 本次没有识别到明显新增的变量关系链。")
    for row in relation_rows:
        lines.append(f"- {row.title}: {row.novel_relations}")

    lines.extend(["", "## 新增文献简表"])
    for row in new_rows[:12]:
        lines.extend(
            [
                f"### {row.title}",
                f"- 来源：{row.source_name}",
                f"- 作者：{row.authors or '待补充'}",
                f"- 年份：{row.year or '待补充'}",
                f"- 聚焦编码：{row.axial_summary}",
                f"- 主轴概括：{row.selective_summary}",
                f"- 关系链：{row.axial_relations or '待补充'}",
                f"- 命题草案：{row.selective_proposition or '待补充'}",
                f"- 编码可信度：{row.coding_confidence or '待补充'}",
                f"- 证据句：{row.evidence_sentences or '待补充'}",
                "",
            ]
        )

    report_path.write_text("\n".join(lines).strip() + "\n", encoding="utf-8")


def main() -> int:
    parser = ChineseArgumentParser(description="扎根理论文献每日监测器")
    parser.add_argument("--config", type=Path, required=True, help="JSON 配置文件路径")
    parser.add_argument("--ask", type=str, default="", help="基于当前文献库提问")
    parser.add_argument("--generate-report", type=str, default="", help="为某个主题生成行业报告")
    parser.add_argument("--skip-monitor", action="store_true", help="跳过重新检索，直接复用现有 literature_table.csv")
    args = parser.parse_args()

    user_config = json.loads(args.config.read_text(encoding="utf-8"))
    config = merge_config(user_config)
    if not config["queries"]:
        print("配置中至少需要一条 queries。", file=sys.stderr)
        return 1

    outdir = Path(config["outdir"]).expanduser().resolve()
    outdir.mkdir(parents=True, exist_ok=True)
    pdf_dir = outdir / "pdfs"
    pdf_dir.mkdir(parents=True, exist_ok=True)

    results_csv = outdir / "literature_table.csv"
    results_xlsx = outdir / "literature_table.xlsx"
    report_path = outdir / f"daily_report_{date.today().isoformat()}.md"
    state_path = outdir / "run_state.json"
    search_history_path = outdir / "search_history.jsonl"
    run_history_path = outdir / "run_history.jsonl"
    download_history_path = outdir / "download_history.json"
    translation_cache_path = outdir / "translation_cache.json"
    theme_memory_path = outdir / "theme_memory.json"
    agent_trace_path = outdir / "agent_trace.jsonl"
    agent_memory_path = outdir / "agent_memory.json"
    compact_context_path = outdir / "compact_context.md"

    baseline_labels, baseline_ready = load_baseline_labels(config)
    baseline_relation_signatures = load_baseline_relation_signatures(config)
    discovered_skills, selected_skills = discover_skills(config, " ".join(config["queries"]))
    agent_plan = build_agent_plan(config, selected_skills)
    write_snapshot(
        config,
        baseline_labels,
        baseline_relation_signatures,
        outdir,
        discovered_skills,
        selected_skills,
        agent_plan,
    )
    log_agent_trace(
        agent_trace_path,
        step="discover_skills",
        status="ok",
        discovered_count=len(discovered_skills),
        selected=[item.get("name", "") for item in selected_skills],
    )

    previous_state = load_json(state_path, {})
    resumed = previous_state.get("status") == "running"
    new_rows: list[MonitorRow] = []
    existing_rows: list[MonitorRow] = []
    if args.skip_monitor:
        existing_rows = load_existing_rows(results_csv)
        if not existing_rows:
            print("未找到现有 literature_table.csv，不能使用 --skip-monitor。", file=sys.stderr)
            return 1
        update_state(
            state_path,
            status="reusing_existing",
            project_name=config["project_name"],
            started_at=datetime.now().isoformat(timespec="seconds"),
            queries=config["queries"],
            sources=config["sources"],
            current_step="reuse_existing_rows",
        )
        log_agent_trace(
            agent_trace_path,
            step="reuse_existing_rows",
            status="ok",
            total_rows=len(existing_rows),
        )
    else:
        update_state(
            state_path,
            status="running",
            project_name=config["project_name"],
            started_at=datetime.now().isoformat(timespec="seconds"),
            queries=config["queries"],
            sources=config["sources"],
        )

        existing_rows = load_existing_rows(results_csv)
        seen_keys = {
            candidate_key(row.title, row.doi, row.external_id)
            for row in existing_rows
            if row.title
        }
        download_history = load_download_history(download_history_path)
        translation_cache = load_translation_cache(translation_cache_path)

        to_date = date.today()
        from_date = to_date - timedelta(days=int(config["days_back"]))

        try:
            for query in config["queries"]:
                for source in config["sources"]:
                    log_agent_trace(
                        agent_trace_path,
                        step=f"search_{source}",
                        status="start",
                        query=query,
                    )
                    update_state(
                        state_path,
                        current_query=query,
                        current_source=source,
                        processed=len(new_rows),
                        current_step=f"search_{source}",
                        selected_skills=[item.get("name", "") for item in selected_skills],
                        agent_plan=agent_plan,
                    )
                    try:
                        candidates = fetch_candidates_for_source(
                            source=source,
                            query=query,
                            config=config,
                            from_date=from_date.isoformat(),
                            to_date=to_date.isoformat(),
                        )
                    except Exception as exc:
                        append_jsonl(
                            search_history_path,
                            {
                                "timestamp": datetime.now().isoformat(timespec="seconds"),
                                "query": query,
                                "source": source,
                                "status": "error",
                                "error": str(exc),
                            },
                        )
                        log_agent_trace(
                            agent_trace_path,
                            step=f"search_{source}",
                            status="error",
                            query=query,
                            error=str(exc),
                        )
                        continue

                    append_jsonl(
                        search_history_path,
                        {
                            "timestamp": datetime.now().isoformat(timespec="seconds"),
                            "query": query,
                            "source": source,
                            "status": "ok",
                            "count": len(candidates),
                        },
                    )
                    log_agent_trace(
                        agent_trace_path,
                        step=f"search_{source}",
                        status="ok",
                        query=query,
                        count=len(candidates),
                    )

                    for candidate in candidates:
                        key = candidate_key(candidate.title, candidate.doi, candidate.external_id)
                        if key in seen_keys:
                            continue
                        row, download_history, translation_cache = make_row(
                            candidate=candidate,
                            config=config,
                            baseline_labels=baseline_labels,
                            baseline_relation_signatures=baseline_relation_signatures,
                            baseline_ready=baseline_ready,
                            pdf_dir=pdf_dir,
                            download_history=download_history,
                            translation_cache=translation_cache,
                        )
                        existing_rows.append(row)
                        new_rows.append(row)
                        seen_keys.add(key)
                        update_state(
                            state_path,
                            current_title=row.title,
                            processed=len(new_rows),
                            total_rows=len(existing_rows),
                            current_step="grounded_coding",
                        )
                        if bool(config.get("agent", {}).get("trace_candidates", False)):
                            log_agent_trace(
                                agent_trace_path,
                                step="candidate_processed",
                                status="ok",
                                title=row.title,
                                source_name=row.source_name,
                                alert_level=row.alert_level,
                            )
                        time.sleep(float(config.get("sleep_seconds", 0.4)))
        except Exception as exc:
            update_state(
                state_path,
                status="failed",
                finished_at=datetime.now().isoformat(timespec="seconds"),
                error=str(exc),
            )
            save_download_history(download_history_path, download_history)
            save_json(translation_cache_path, translation_cache)
            raise

        existing_rows.sort(key=lambda row: (row.publication_date, row.year, row.title), reverse=True)
        write_csv(existing_rows, results_csv)
        write_xlsx(existing_rows, results_xlsx)
        save_download_history(download_history_path, download_history)
        save_json(translation_cache_path, translation_cache)

    theme_memory = write_theme_memory(existing_rows, theme_memory_path)
    log_agent_trace(
        agent_trace_path,
        step="memory_persist",
        status="ok",
        total_rows=len(existing_rows),
        new_rows=len(new_rows),
    )
    agent_memory = write_agent_memory(
        path=agent_memory_path,
        config=config,
        rows=existing_rows,
        new_rows=new_rows,
        theme_memory=theme_memory,
        selected_skills=selected_skills,
    )
    compact_context_written = write_compact_context(
        path=compact_context_path,
        config=config,
        rows=existing_rows,
        new_rows=new_rows,
        theme_memory=theme_memory,
        agent_memory=agent_memory,
        selected_skills=selected_skills,
    )
    log_agent_trace(
        agent_trace_path,
        step="context_compression",
        status="ok",
        compact_context=compact_context_written,
    )
    if not args.skip_monitor:
        write_daily_report(
            rows=existing_rows,
            new_rows=new_rows,
            baseline_labels=baseline_labels,
            baseline_ready=baseline_ready,
            report_path=report_path,
            config=config,
            theme_memory=theme_memory,
            resumed=resumed,
            selected_skills=selected_skills,
            compact_context_path=compact_context_written,
            agent_memory_path=str(agent_memory_path),
        )

    qa_output = ""
    industry_report_output = ""
    if args.ask.strip():
        qa_output = answer_question(
            question=args.ask.strip(),
            rows=existing_rows,
            config=config,
            outdir=outdir,
            agent_trace_path=agent_trace_path,
        )
    if args.generate_report.strip():
        industry_report_output = generate_industry_report(
            topic=args.generate_report.strip(),
            rows=existing_rows,
            config=config,
            outdir=outdir,
            theme_memory=theme_memory,
            agent_trace_path=agent_trace_path,
        )

    summary = {
        "project_name": config["project_name"],
        "new_rows": len(new_rows),
        "total_rows": len(existing_rows),
        "results_csv": str(results_csv),
        "results_xlsx": str(results_xlsx),
        "daily_report": str(report_path) if report_path.exists() else "",
        "pdf_dir": str(pdf_dir),
        "state_file": str(state_path),
        "search_history": str(search_history_path),
        "run_history": str(run_history_path),
        "download_history": str(download_history_path),
        "theme_memory": str(theme_memory_path),
        "agent_trace": str(agent_trace_path),
        "agent_memory": str(agent_memory_path),
        "compact_context": str(compact_context_path),
        "qa_answer": qa_output,
        "industry_report": industry_report_output,
    }
    append_jsonl(
        run_history_path,
        {
            "timestamp": datetime.now().isoformat(timespec="seconds"),
            "project_name": config["project_name"],
            "new_rows": len(new_rows),
            "total_rows": len(existing_rows),
            "sources": config["sources"],
            "queries": config["queries"],
            "skip_monitor": bool(args.skip_monitor),
            "qa_output": qa_output,
            "industry_report_output": industry_report_output,
        },
    )
    update_state(
        state_path,
        status="completed",
        finished_at=datetime.now().isoformat(timespec="seconds"),
        summary=summary,
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
